#!/usr/bin/env python3
"""
Enhanced P&L Report Generator

Generates comprehensive Profit & Loss statements matching cloud_kitchen_pl.xlsx format.
Includes detailed COGS, OPEX, depreciation, and KPIs.

Usage:
    python generate_pnl.py <start_date> <end_date> [format] [restaurant_id]

Examples:
    python generate_pnl.py 2024-01-01 2024-01-31
    python generate_pnl.py 2024-01-01 2024-01-31 excel
"""

import os
import sys
from pathlib import Path
from datetime import datetime, timedelta
from pymongo import MongoClient
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def get_mongodb_connection():
    """Get MongoDB connection"""
    mongodb_uri = os.getenv('MONGODB_URI', 'mongodb://localhost:27017')
    db_name = os.getenv('DB_NAME', 'ahar_pos')

    try:
        client = MongoClient(mongodb_uri, serverSelectionTimeoutMS=5000)
        client.server_info()
        db = client[db_name]
        return db
    except Exception as e:
        print(f"ERROR:Failed to connect to MongoDB: {e}")
        sys.exit(1)


def parse_date_range(start_date: str, end_date: str):
    """Parse and validate date range"""
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)

        num_days = (end_dt - start_dt).days + 1
        return start_dt, end_dt, num_days
    except ValueError as e:
        print(f"ERROR:Invalid date format. Use YYYY-MM-DD: {e}")
        sys.exit(1)


def clamp_date_range_to_available_orders(db, start_dt: datetime, end_dt: datetime):
    """
    Clamp date range to available data in `orders.created_at`.

    This prevents empty P&L when the requested range extends beyond the dataset's
    latest `created_at` (common with static/seeded datasets).
    """
    try:
        latest = db.orders.find({"created_at": {"$exists": True}}).sort("created_at", -1).limit(1)
        latest_doc = next(latest, None)
        latest_created_at = latest_doc.get("created_at") if latest_doc else None
        if not isinstance(latest_created_at, datetime):
            return start_dt, end_dt, False

        adjusted = False

        if start_dt > latest_created_at:
            start_dt = latest_created_at.replace(hour=0, minute=0, second=0, microsecond=0)
            adjusted = True

        if end_dt > latest_created_at:
            end_dt = latest_created_at.replace(hour=23, minute=59, second=59, microsecond=999999)
            adjusted = True

        if end_dt < start_dt:
            end_dt = start_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
            adjusted = True

        return start_dt, end_dt, adjusted
    except Exception:
        return start_dt, end_dt, False


def get_restaurant_settings(db, restaurant_id: str):
    """Get restaurant settings or use defaults"""
    settings = db.restaurant_settings.find_one({"restaurant_id": restaurant_id})
    if not settings:
        print(f"WARNING:No settings found for '{restaurant_id}', using defaults")
        # Return minimal defaults
        return {
            "platform_settings": {
                "zomato_commission_rate": 0.23,
                "swiggy_commission_rate": 0.23,
                "gst_rate": 0.05,
                "cancellation_rate": 0.015
            },
            "role_salaries": {},
            "pf_esic_settings": {"pf_employer_rate": 0.12, "esic_employer_rate": 0.0175},
            "overtime_settings": {},
            "occupancy_costs": {},
            "technology_costs": {},
            "marketing_budgets": {},
            "general_admin_costs": {},
            "depreciation_amortization": {"equipment_depreciation": 1500000, "brand_amortization": 200000},
            "finance_costs": {"loan_interest": 800000, "bank_charges": 150000},
            "tax_settings": {"presumptive_tax_rate": 0.26}
        }
    return settings


def calculate_revenue_breakdown(db, restaurant_id: str, start_dt, end_dt, settings):
    """Calculate detailed revenue breakdown by category"""

    # Get delivery orders
    delivery_query = {"order_date": {"$gte": start_dt, "$lte": end_dt}}
    sample_delivery = db.delivery_orders.find_one({})
    if sample_delivery and "restaurant_id" in sample_delivery:
        # Only filter if the requested restaurant_id exists in data
        if db.delivery_orders.count_documents({"restaurant_id": restaurant_id}, limit=1) > 0:
            delivery_query["restaurant_id"] = restaurant_id
    delivery_orders = list(db.delivery_orders.find(delivery_query))

    # Get dine-in/takeaway orders (items are embedded)
    # Note: orders may not have restaurant_id field, so make it optional
    order_query = {
        "created_at": {"$gte": start_dt, "$lte": end_dt},
        "status": {"$in": ["COMPLETED", "completed", "sent_to_kitchen", "in_progress"]}
    }
    # Only filter by restaurant_id if it exists AND matches data
    sample_order = db.orders.find_one({})
    if sample_order and "restaurant_id" in sample_order:
        if db.orders.count_documents({"restaurant_id": restaurant_id}, limit=1) > 0:
            order_query["restaurant_id"] = restaurant_id

    regular_orders = list(db.orders.find(order_query))

    # Get menu items for categorization - index by menu_item_id string
    menu_items = {m.get("menu_item_id", str(m["_id"])): m for m in db.menu_items.find({})}

    # Initialize revenue trackers
    revenue = {
        "veg": 0,
        "non_veg": 0,
        "sides": 0,
        "beverages": 0,
        "desserts": 0,
        "total_orders": 0
    }

    channel_revenue = defaultdict(float)

    # Process delivery orders (only have totals, no item breakdown)
    for order in delivery_orders:
        channel = order.get("order_channel", "WalkIn")
        total = order.get("total_inr", 0)
        channel_revenue[channel] += total
        revenue["total_orders"] += 1

    # Process regular orders with embedded items for category breakdown
    for order in regular_orders:
        items = order.get("items", [])
        revenue["total_orders"] += 1

        for item in items:
            menu_id = item.get("menu_item_id")
            menu = menu_items.get(menu_id, {})
            category = menu.get("category", "")
            price = item.get("price_snapshot", 0)
            quantity = item.get("quantity", 1)
            item_total = price * quantity / 100  # Convert paise to rupees

            # Categorize based on menu category and tags
            if "Beverage" in category or "Drink" in category:
                revenue["beverages"] += item_total
            elif "Dessert" in category or "Sweet" in category:
                revenue["desserts"] += item_total
            elif "Starter" in category or "Appetizer" in category or "Side" in category:
                revenue["sides"] += item_total
            else:
                # Main course - check if veg or non-veg based on tags
                tags = menu.get("tags", [])
                tags_upper = [tag.upper() for tag in tags]
                if any(tag in tags_upper for tag in ["CHICKEN", "FISH", "MEAT", "PORK", "LAMB", "BEEF", "EGG"]):
                    revenue["non_veg"] += item_total
                else:
                    revenue["veg"] += item_total

    # Calculate GMV
    gross_gmv = sum([order.get("total_inr", 0) for order in delivery_orders]) / 100
    for order in regular_orders:
        gross_gmv += order.get("total_amount", 0) / 100

    # Apply cancellation rate
    cancellation_rate = settings["platform_settings"]["cancellation_rate"]
    cancellations = gross_gmv * cancellation_rate
    net_gmv = gross_gmv - cancellations

    # Calculate platform commissions
    zomato_gmv = channel_revenue.get("Zomato", 0) / 100
    swiggy_gmv = channel_revenue.get("Swiggy", 0) / 100

    zomato_commission = zomato_gmv * settings["platform_settings"]["zomato_commission_rate"]
    swiggy_commission = swiggy_gmv * settings["platform_settings"]["swiggy_commission_rate"]

    # Calculate GST
    gst = net_gmv * settings["platform_settings"]["gst_rate"]

    # Net revenue
    net_revenue = net_gmv - zomato_commission - swiggy_commission - gst

    return {
        "gross_gmv": gross_gmv,
        "veg_revenue": revenue["veg"],
        "non_veg_revenue": revenue["non_veg"],
        "sides_revenue": revenue["sides"],
        "beverages_revenue": revenue["beverages"],
        "desserts_revenue": revenue["desserts"],
        "cancellations": cancellations,
        "net_gmv": net_gmv,
        "zomato_commission": zomato_commission,
        "swiggy_commission": swiggy_commission,
        "gst": gst,
        "net_revenue": net_revenue,
        "total_orders": revenue["total_orders"],
        "channel_revenue": {k: v/100 for k, v in channel_revenue.items()}
    }


def calculate_cogs(db, restaurant_id: str, start_dt, end_dt):
    """Calculate Cost of Goods Sold with detailed breakdown"""

    # Get all completed orders in date range (items are embedded)
    # Note: orders may not have restaurant_id field, so make it optional
    order_query = {
        "created_at": {"$gte": start_dt, "$lte": end_dt},
        "status": {"$in": ["COMPLETED", "completed", "sent_to_kitchen", "in_progress"]}
    }
    # Only filter by restaurant_id if it exists AND matches data
    sample_order = db.orders.find_one({})
    if sample_order and "restaurant_id" in sample_order:
        if db.orders.count_documents({"restaurant_id": restaurant_id}, limit=1) > 0:
            order_query["restaurant_id"] = restaurant_id

    regular_orders = list(db.orders.find(order_query))

    # Get recipe BOM - index by menu_item_id string
    recipe_bom = {}
    for bom in db.recipe_bom.find({}):
        menu_id = bom["menu_item_id"]
        if menu_id not in recipe_bom:
            recipe_bom[menu_id] = []
        # Store ingredients array
        ingredients = bom.get("ingredients", [])
        for ingredient in ingredients:
            recipe_bom[menu_id].append(ingredient)

    # Get material costs
    materials = {m["material_id"]: m for m in db.raw_material_inventory.find({})}

    # Get packaging BOM - handle both string and ObjectId formats
    packaging_bom = {}
    menu_items = {m.get("menu_item_id"): str(m["_id"]) for m in db.menu_items.find({})}

    for bom in db.packaging_bom.find({}):
        menu_id = bom["menu_item_id"]
        if menu_id not in packaging_bom:
            packaging_bom[menu_id] = []
        packaging_bom[menu_id].append(bom)

        # Also index by string ID if menu_id is ObjectId format
        for string_id, obj_id in menu_items.items():
            if menu_id == obj_id and string_id not in packaging_bom:
                packaging_bom[string_id] = packaging_bom[menu_id]

    # Get packaging costs
    packaging_materials = {p["packaging_id"]: p for p in db.packaging_materials.find({})}

    # Calculate raw material costs by category
    material_costs = defaultdict(float)
    packaging_costs = defaultdict(float)

    # Process embedded items from orders
    for order in regular_orders:
        items = order.get("items", [])

        for item in items:
            menu_id = item.get("menu_item_id")
            quantity = item.get("quantity", 1)

            # Calculate raw material cost from recipe BOM
            if menu_id in recipe_bom:
                for ingredient in recipe_bom[menu_id]:
                    material_id = ingredient.get("material_id")
                    if material_id in materials:
                        material = materials[material_id]
                        qty_needed = ingredient.get("quantity_per_serving", 0) * quantity
                        unit_cost = material.get("unit_cost_inr", 0)
                        cost = (qty_needed * unit_cost) / 100  # Convert paise to rupees

                        category = material.get("category", "Other")
                        material_costs[category] += cost

            # Calculate packaging cost from packaging BOM
            if menu_id in packaging_bom:
                for pkg_item in packaging_bom[menu_id]:
                    pkg_id = pkg_item.get("packaging_material_id")
                    if pkg_id in packaging_materials:
                        pkg = packaging_materials[pkg_id]
                        qty_needed = pkg_item.get("quantity_per_serving", 1) * quantity
                        unit_cost = pkg.get("unit_cost_inr", 0)
                        cost = (qty_needed * unit_cost) / 100

                        pkg_category = pkg.get("category", "OTHER")
                        packaging_costs[pkg_category] += cost

    # Get wastage and staff meals from stock movements
    wastage = 0
    staff_meals = 0
    qc_sampling = 0

    movements = db.stock_movement_log.find({
        "movement_date": {"$gte": start_dt, "$lte": end_dt},
        "movement_type": {"$in": ["WASTE", "STAFF_MEAL", "QC_SAMPLE"]}
    })

    for mov in movements:
        material_id = mov.get("material_id")
        if material_id in materials:
            material = materials[material_id]
            qty = abs(mov.get("quantity", 0))
            unit_cost = material.get("unit_cost_inr", 0)
            cost = (qty * unit_cost) / 100

            mov_type = mov.get("movement_type")
            if mov_type == "WASTE":
                wastage += cost
            elif mov_type == "STAFF_MEAL":
                staff_meals += cost
            elif mov_type == "QC_SAMPLE":
                qc_sampling += cost

    return {
        "raw_material_by_category": dict(material_costs),
        "total_raw_material": sum(material_costs.values()),
        "packaging_by_category": dict(packaging_costs),
        "total_packaging": sum(packaging_costs.values()),
        "wastage": wastage,
        "staff_meals": staff_meals,
        "qc_sampling": qc_sampling,
        "total_wastage_other": wastage + staff_meals + qc_sampling,
        "total_cogs": sum(material_costs.values()) + sum(packaging_costs.values()) + wastage + staff_meals + qc_sampling
    }


def calculate_labour_costs(db, settings, num_days):
    """Calculate labour costs from user database"""

    # Get active users
    users = list(db.users.find({"status": {"$ne": "inactive"}}))

    role_salaries = settings.get("role_salaries", {})
    overtime_settings = settings.get("overtime_settings", {})
    pf_esic = settings.get("pf_esic_settings", {})

    monthly_factor = num_days / 30

    total_salaries = 0
    role_counts = defaultdict(int)

    for user in users:
        role = user.get("role", "waiter").lower()
        role_counts[role] += 1

        # Get salary for this role (in paise)
        salary = role_salaries.get(role, 2000000)  # Default ₹20,000
        total_salaries += salary

    # Convert to rupees and apply monthly factor
    total_salaries_inr = (total_salaries / 100) * monthly_factor

    # Calculate PF/ESIC
    pf_rate = pf_esic.get("pf_employer_rate", 0.12)
    esic_rate = pf_esic.get("esic_employer_rate", 0.0175)
    total_contribution_rate = pf_rate + esic_rate

    pf_esic_amount = total_salaries_inr * total_contribution_rate

    # Calculate overtime
    total_overtime = 0
    for role, count in role_counts.items():
        overtime_per_role = overtime_settings.get(role, 0)  # in paise
        total_overtime += (overtime_per_role / 100) * count * monthly_factor

    return {
        "salaries": total_salaries_inr,
        "pf_esic": pf_esic_amount,
        "overtime": total_overtime,
        "total_labour": total_salaries_inr + pf_esic_amount + total_overtime,
        "employee_count": len(users)
    }


def calculate_opex(settings, num_days):
    """Calculate operating expenses (proportionate to period)"""

    monthly_factor = num_days / 30

    # E2: Occupancy
    occupancy = settings.get("occupancy_costs", {})
    occupancy_total = sum([
        occupancy.get("rent", 0),
        occupancy.get("electricity", 0),
        occupancy.get("water", 0),
        occupancy.get("internet", 0)
    ]) / 100 * monthly_factor

    # E3: Technology
    technology = settings.get("technology_costs", {})
    technology_total = sum([
        technology.get("pos_software", 0),
        technology.get("platform_subscriptions", 0),
        technology.get("menu_photography_amortized", 0)
    ]) / 100 * monthly_factor

    # E4: Marketing
    marketing = settings.get("marketing_budgets", {})
    marketing_total = sum([
        marketing.get("zomato_ads", 0),
        marketing.get("swiggy_ads", 0),
        marketing.get("social_media", 0),
        marketing.get("influencer", 0),
        marketing.get("self_funded_discounts", 0)
    ]) / 100 * monthly_factor

    # E5: General & Admin
    admin = settings.get("general_admin_costs", {})
    admin_total = sum([
        admin.get("accounting", 0),
        admin.get("legal_compliance", 0),
        admin.get("insurance", 0),
        admin.get("cleaning_supplies", 0),
        admin.get("pest_control", 0),
        admin.get("repairs_maintenance", 0),
        admin.get("gas_lpg", 0),
        admin.get("office_supplies", 0),
        admin.get("miscellaneous", 0)
    ]) / 100 * monthly_factor

    return {
        "occupancy": occupancy_total,
        "technology": technology_total,
        "marketing": marketing_total,
        "general_admin": admin_total,
        "occupancy_details": {k: v/100*monthly_factor for k, v in occupancy.items()},
        "technology_details": {k: v/100*monthly_factor for k, v in technology.items()},
        "marketing_details": {k: v/100*monthly_factor for k, v in marketing.items()},
        "admin_details": {k: v/100*monthly_factor for k, v in admin.items()}
    }


def calculate_depreciation(settings, num_days):
    """Calculate depreciation and amortization"""
    monthly_factor = num_days / 30

    dep_amort = settings.get("depreciation_amortization", {})
    equipment = dep_amort.get("equipment_depreciation", 0) / 100 * monthly_factor
    brand = dep_amort.get("brand_amortization", 0) / 100 * monthly_factor

    return {
        "equipment": equipment,
        "brand": brand,
        "total": equipment + brand
    }


def calculate_finance_costs(settings, num_days):
    """Calculate finance costs"""
    monthly_factor = num_days / 30

    finance = settings.get("finance_costs", {})
    interest = finance.get("loan_interest", 0) / 100 * monthly_factor
    bank_charges = finance.get("bank_charges", 0) / 100 * monthly_factor

    return {
        "interest": interest,
        "bank_charges": bank_charges,
        "total": interest + bank_charges
    }


def calculate_tax(pbt, settings):
    """Calculate income tax"""
    if pbt <= 0:
        return 0

    tax_rate = settings.get("tax_settings", {}).get("presumptive_tax_rate", 0.26)
    return pbt * tax_rate


def format_currency(amount: float) -> str:
    """Format amount as Indian Rupee currency"""
    return f"₹{amount:,.2f}"


def generate_text_report(
    metrics: dict,
    start_date: str,
    end_date: str,
    period_note: str = "",
) -> str:
    """Generate detailed P&L as formatted text"""

    lines = []
    lines.append("=" * 70)
    lines.append("PROFIT & LOSS STATEMENT")
    lines.append(f"Period: {start_date} to {end_date}")
    lines.append("=" * 70)
    lines.append("")
    if period_note:
        lines.append(period_note)
        lines.append("")

    # A. GMV
    lines.append("A. GROSS MERCHANDISE VALUE (GMV)")
    lines.append("-" * 70)
    if metrics["revenue"]["veg_revenue"] > 0:
        lines.append(f"  Main Course Revenue – Vegetarian           {format_currency(metrics['revenue']['veg_revenue']):>20}")
    if metrics["revenue"]["non_veg_revenue"] > 0:
        lines.append(f"  Main Course Revenue – Non-Vegetarian       {format_currency(metrics['revenue']['non_veg_revenue']):>20}")
    if metrics["revenue"]["sides_revenue"] > 0:
        lines.append(f"  Sides Revenue                              {format_currency(metrics['revenue']['sides_revenue']):>20}")
    if metrics["revenue"]["beverages_revenue"] > 0:
        lines.append(f"  Beverages Revenue                          {format_currency(metrics['revenue']['beverages_revenue']):>20}")
    if metrics["revenue"]["desserts_revenue"] > 0:
        lines.append(f"  Desserts Revenue                           {format_currency(metrics['revenue']['desserts_revenue']):>20}")
    lines.append(f"  Gross GMV                                  {format_currency(metrics['revenue']['gross_gmv']):>20}")
    lines.append(f"  Less: Cancellations (1.5%)                 {format_currency(-metrics['revenue']['cancellations']):>20}")
    lines.append("-" * 70)
    lines.append(f"  Net GMV                                    {format_currency(metrics['revenue']['net_gmv']):>20}")
    lines.append("")

    # B. Revenue
    lines.append("B. REVENUE (Net Platform Commissions & GST)")
    lines.append("-" * 70)
    lines.append(f"  Less: Zomato Commission (23%)      {format_currency(-metrics['revenue']['zomato_commission']):>30}")
    lines.append(f"  Less: Swiggy Commission (23%)      {format_currency(-metrics['revenue']['swiggy_commission']):>30}")
    lines.append(f"  Less: GST on Food (5%)             {format_currency(-metrics['revenue']['gst']):>30}")
    lines.append("-" * 70)
    lines.append(f"  NET REVENUE                        {format_currency(metrics['revenue']['net_revenue']):>30}")
    lines.append("")

    # C. COGS
    lines.append("C. COST OF GOODS SOLD (COGS)")
    lines.append("-" * 70)
    lines.append("  C1. Raw Material – Food & Beverage")
    # Sort categories in preferred order
    category_order = ["Proteins", "Vegetables", "Dairy", "Bakery", "Spices", "Oils", "Beverages"]
    for category in category_order:
        if category in metrics["cogs"]["raw_material_by_category"]:
            amount = metrics["cogs"]["raw_material_by_category"][category]
            lines.append(f"    {category:<38} {format_currency(amount):>20}")
    # Add any remaining categories not in the order
    for category, amount in sorted(metrics["cogs"]["raw_material_by_category"].items()):
        if category not in category_order:
            lines.append(f"    {category:<38} {format_currency(amount):>20}")
    lines.append(f"    Sub-Total: Raw Material                {format_currency(metrics['cogs']['total_raw_material']):>20}")
    lines.append("")
    lines.append("  C2. Packaging Material")
    # Sort packaging categories
    pkg_order = ["LABELS", "PRIMARY", "SECONDARY"]
    for pkg_cat in pkg_order:
        if pkg_cat in metrics["cogs"]["packaging_by_category"]:
            amount = metrics["cogs"]["packaging_by_category"][pkg_cat]
            cat_name = pkg_cat.title() + " Packaging"
            lines.append(f"    {cat_name:<38} {format_currency(amount):>20}")
    # Add any remaining categories
    for category, amount in sorted(metrics["cogs"]["packaging_by_category"].items()):
        if category not in pkg_order:
            cat_name = category.title() + " Packaging"
            lines.append(f"    {cat_name:<38} {format_currency(amount):>20}")
    lines.append(f"    Sub-Total: Packaging                   {format_currency(metrics['cogs']['total_packaging']):>20}")
    lines.append("")
    lines.append("  C3. Wastage & Other Food Costs")
    lines.append(f"    Wastage & Spoilage                     {format_currency(metrics['cogs']['wastage']):>20}")
    lines.append(f"    Staff Meals                            {format_currency(metrics['cogs']['staff_meals']):>20}")
    lines.append(f"    Quality Control                        {format_currency(metrics['cogs']['qc_sampling']):>20}")
    lines.append(f"    Sub-Total: Wastage & Other             {format_currency(metrics['cogs']['total_wastage_other']):>20}")
    lines.append("-" * 70)
    lines.append(f"  TOTAL COGS                                 {format_currency(metrics['cogs']['total_cogs']):>20}")
    lines.append("")

    # D. Gross Profit
    gross_profit = metrics["revenue"]["net_revenue"] - metrics["cogs"]["total_cogs"]
    gross_margin = (gross_profit / metrics["revenue"]["net_revenue"] * 100) if metrics["revenue"]["net_revenue"] > 0 else 0
    lines.append("D. GROSS PROFIT")
    lines.append("-" * 70)
    lines.append(f"  Gross Profit                       {format_currency(gross_profit):>30}")
    lines.append(f"  Gross Margin %                     {gross_margin:>29.1f}%")
    lines.append("")

    # E. OPEX
    lines.append("E. OPERATING EXPENSES (OPEX)")
    lines.append("-" * 70)
    lines.append("  E1. Labour & HR Costs")
    lines.append(f"    Staff Salaries ({metrics['labour']['employee_count']} employees)      {format_currency(metrics['labour']['salaries']):>30}")
    lines.append(f"    PF/ESIC (13.75%)               {format_currency(metrics['labour']['pf_esic']):>30}")
    lines.append(f"    Overtime Allowance             {format_currency(metrics['labour']['overtime']):>30}")
    lines.append(f"    Sub-Total: Labour              {format_currency(metrics['labour']['total_labour']):>30}")
    lines.append("")
    lines.append("  E2. Occupancy & Utilities")
    for key, value in metrics["opex"]["occupancy_details"].items():
        lines.append(f"    {key.replace('_', ' ').title():<30} {format_currency(value):>30}")
    lines.append(f"    Sub-Total: Occupancy           {format_currency(metrics['opex']['occupancy']):>30}")
    lines.append("")
    lines.append("  E3. Technology & Software")
    for key, value in metrics["opex"]["technology_details"].items():
        lines.append(f"    {key.replace('_', ' ').title():<30} {format_currency(value):>30}")
    lines.append(f"    Sub-Total: Technology          {format_currency(metrics['opex']['technology']):>30}")
    lines.append("")
    lines.append("  E4. Sales & Marketing")
    for key, value in metrics["opex"]["marketing_details"].items():
        lines.append(f"    {key.replace('_', ' ').title():<30} {format_currency(value):>30}")
    lines.append(f"    Sub-Total: Marketing           {format_currency(metrics['opex']['marketing']):>30}")
    lines.append("")
    lines.append("  E5. General & Administrative")
    for key, value in metrics["opex"]["admin_details"].items():
        lines.append(f"    {key.replace('_', ' ').title():<30} {format_currency(value):>30}")
    lines.append(f"    Sub-Total: General & Admin     {format_currency(metrics['opex']['general_admin']):>30}")
    lines.append("-" * 70)
    total_opex = (metrics['labour']['total_labour'] + metrics['opex']['occupancy'] +
                  metrics['opex']['technology'] + metrics['opex']['marketing'] +
                  metrics['opex']['general_admin'])
    lines.append(f"  TOTAL OPEX                         {format_currency(total_opex):>30}")
    lines.append("")

    # F. EBITDA
    ebitda = gross_profit - total_opex
    ebitda_margin = (ebitda / metrics["revenue"]["net_revenue"] * 100) if metrics["revenue"]["net_revenue"] > 0 else 0
    lines.append("F. EBITDA")
    lines.append("-" * 70)
    lines.append(f"  EBITDA                             {format_currency(ebitda):>30}")
    lines.append(f"  EBITDA Margin %                    {ebitda_margin:>29.1f}%")
    lines.append("")

    # G. Depreciation
    lines.append("G. DEPRECIATION & AMORTIZATION")
    lines.append("-" * 70)
    lines.append(f"  Equipment Depreciation             {format_currency(metrics['depreciation']['equipment']):>30}")
    lines.append(f"  Brand Amortization                 {format_currency(metrics['depreciation']['brand']):>30}")
    lines.append(f"  Total D&A                          {format_currency(metrics['depreciation']['total']):>30}")
    lines.append("")

    # H. EBIT
    ebit = ebitda - metrics['depreciation']['total']
    lines.append("H. EBIT")
    lines.append("-" * 70)
    lines.append(f"  EBIT                               {format_currency(ebit):>30}")
    lines.append("")

    # I. Finance Costs
    lines.append("I. FINANCE COSTS")
    lines.append("-" * 70)
    lines.append(f"  Interest on Loans                  {format_currency(metrics['finance']['interest']):>30}")
    lines.append(f"  Bank Charges                       {format_currency(metrics['finance']['bank_charges']):>30}")
    lines.append(f"  Total Finance Costs                {format_currency(metrics['finance']['total']):>30}")
    lines.append("")

    # J. PBT
    pbt = ebit - metrics['finance']['total']
    lines.append("J. PROFIT BEFORE TAX (PBT)")
    lines.append("-" * 70)
    lines.append(f"  PBT                                {format_currency(pbt):>30}")
    lines.append("")

    # K. Tax
    lines.append("K. INCOME TAX")
    lines.append("-" * 70)
    lines.append(f"  Income Tax (26%)                   {format_currency(metrics['tax']):>30}")
    lines.append("")

    # L. PAT
    pat = pbt - metrics['tax']
    pat_margin = (pat / metrics["revenue"]["net_revenue"] * 100) if metrics["revenue"]["net_revenue"] > 0 else 0
    lines.append("=" * 70)
    lines.append("L. PROFIT AFTER TAX (PAT) ★")
    lines.append("=" * 70)
    lines.append(f"  PAT                                {format_currency(pat):>30}")
    lines.append(f"  PAT Margin %                       {pat_margin:>29.1f}%")
    lines.append("=" * 70)
    lines.append("")

    # M. KPIs
    lines.append("M. KEY PERFORMANCE INDICATORS")
    lines.append("-" * 70)
    lines.append(f"  Total Orders                       {metrics['revenue']['total_orders']:>30,}")
    aov = metrics["revenue"]["net_gmv"] / metrics["revenue"]["total_orders"] if metrics["revenue"]["total_orders"] > 0 else 0
    lines.append(f"  Average Order Value                {format_currency(aov):>30}")
    lines.append(f"  Food Cost % (of Net Revenue)       {(metrics['cogs']['total_raw_material']/metrics['revenue']['net_revenue']*100) if metrics['revenue']['net_revenue'] > 0 else 0:>29.1f}%")
    lines.append(f"  Labour Cost % (of Net Revenue)     {(metrics['labour']['total_labour']/metrics['revenue']['net_revenue']*100) if metrics['revenue']['net_revenue'] > 0 else 0:>29.1f}%")
    total_commission = metrics['revenue']['zomato_commission'] + metrics['revenue']['swiggy_commission']
    lines.append(f"  Platform Commission % (of GMV)     {(total_commission/metrics['revenue']['gross_gmv']*100) if metrics['revenue']['gross_gmv'] > 0 else 0:>29.1f}%")
    lines.append("=" * 70)

    return "\n".join(lines)


def generate_excel_report(metrics: dict, start_date: str, end_date: str, output_path: str) -> str:
    """
    Generate P&L as Excel file with formatting

    Args:
        metrics: Calculated P&L metrics
        start_date: Start date string
        end_date: End date string
        output_path: Path to save Excel file

    Returns:
        Path to saved Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Statement"

    # Styles
    header_font = Font(name='Arial', size=14, bold=True)
    section_font = Font(name='Arial', size=12, bold=True)
    subsection_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subsection_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Header
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "PROFIT & LOSS STATEMENT"
    cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = center_align
    row += 1

    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = f"Period: {start_date} to {end_date}"
    cell.font = subsection_font
    cell.alignment = center_align
    row += 2

    # Helper function to add section
    def add_section(title, items, row_num):
        ws.merge_cells(f'A{row_num}:C{row_num}')
        cell = ws[f'A{row_num}']
        cell.value = title
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = left_align
        row_num += 1

        for item in items:
            ws[f'A{row_num}'] = item['label']
            ws[f'A{row_num}'].font = bold_font if item.get('bold') else normal_font
            ws[f'A{row_num}'].alignment = left_align

            if item.get('indent'):
                ws[f'A{row_num}'].alignment = Alignment(horizontal='left', indent=item['indent'])

            if 'value' in item:
                ws[f'C{row_num}'] = item['value']
                ws[f'C{row_num}'].font = bold_font if item.get('bold') else normal_font
                ws[f'C{row_num}'].alignment = right_align
                ws[f'C{row_num}'].number_format = '#,##0.00' if isinstance(item['value'], (int, float)) else '@'

            if item.get('fill'):
                ws[f'A{row_num}'].fill = item['fill']
                ws[f'C{row_num}'].fill = item['fill']

            row_num += 1

        row_num += 1
        return row_num

    # A. GMV
    gmv_items = []
    if metrics["revenue"]["veg_revenue"] > 0:
        gmv_items.append({'label': 'Main Course Revenue – Vegetarian', 'value': metrics["revenue"]["veg_revenue"], 'indent': 1})
    if metrics["revenue"]["non_veg_revenue"] > 0:
        gmv_items.append({'label': 'Main Course Revenue – Non-Vegetarian', 'value': metrics["revenue"]["non_veg_revenue"], 'indent': 1})
    if metrics["revenue"]["sides_revenue"] > 0:
        gmv_items.append({'label': 'Sides Revenue', 'value': metrics["revenue"]["sides_revenue"], 'indent': 1})
    if metrics["revenue"]["beverages_revenue"] > 0:
        gmv_items.append({'label': 'Beverages Revenue', 'value': metrics["revenue"]["beverages_revenue"], 'indent': 1})
    if metrics["revenue"]["desserts_revenue"] > 0:
        gmv_items.append({'label': 'Desserts Revenue', 'value': metrics["revenue"]["desserts_revenue"], 'indent': 1})

    gmv_items.extend([
        {'label': 'Gross GMV', 'value': metrics["revenue"]["gross_gmv"], 'indent': 1, 'bold': True},
        {'label': 'Less: Cancellations (1.5%)', 'value': -metrics["revenue"]["cancellations"], 'indent': 1},
        {'label': 'Net GMV', 'value': metrics["revenue"]["net_gmv"], 'bold': True, 'fill': subsection_fill}
    ])
    row = add_section("A. GROSS MERCHANDISE VALUE (GMV)", gmv_items, row)

    # B. Revenue
    revenue_items = [
        {'label': 'Less: Zomato Commission (23%)', 'value': -metrics["revenue"]["zomato_commission"], 'indent': 1},
        {'label': 'Less: Swiggy Commission (23%)', 'value': -metrics["revenue"]["swiggy_commission"], 'indent': 1},
        {'label': 'Less: GST on Food (5%)', 'value': -metrics["revenue"]["gst"], 'indent': 1},
        {'label': 'NET REVENUE', 'value': metrics["revenue"]["net_revenue"], 'bold': True, 'fill': total_fill}
    ]
    row = add_section("B. REVENUE (Net Platform Commissions & GST)", revenue_items, row)

    # C. COGS
    cogs_items = [{'label': 'C1. Raw Material – Food & Beverage', 'bold': True}]
    category_order = ["Proteins", "Vegetables", "Dairy", "Bakery", "Spices", "Oils", "Beverages"]
    for category in category_order:
        if category in metrics["cogs"]["raw_material_by_category"]:
            cogs_items.append({'label': category, 'value': metrics["cogs"]["raw_material_by_category"][category], 'indent': 2})
    cogs_items.append({'label': 'Sub-Total: Raw Material', 'value': metrics["cogs"]["total_raw_material"], 'indent': 2, 'bold': True})

    cogs_items.append({'label': ''})
    cogs_items.append({'label': 'C2. Packaging Material', 'bold': True})
    pkg_order = ["LABELS", "PRIMARY", "SECONDARY"]
    for pkg_cat in pkg_order:
        if pkg_cat in metrics["cogs"]["packaging_by_category"]:
            label = pkg_cat.title() + " Packaging"
            cogs_items.append({'label': label, 'value': metrics["cogs"]["packaging_by_category"][pkg_cat], 'indent': 2})
    cogs_items.append({'label': 'Sub-Total: Packaging', 'value': metrics["cogs"]["total_packaging"], 'indent': 2, 'bold': True})

    cogs_items.append({'label': ''})
    cogs_items.append({'label': 'C3. Wastage & Other Food Costs', 'bold': True})
    cogs_items.extend([
        {'label': 'Wastage & Spoilage', 'value': metrics["cogs"]["wastage"], 'indent': 2},
        {'label': 'Staff Meals', 'value': metrics["cogs"]["staff_meals"], 'indent': 2},
        {'label': 'Quality Control', 'value': metrics["cogs"]["qc_sampling"], 'indent': 2},
        {'label': 'Sub-Total: Wastage & Other', 'value': metrics["cogs"]["total_wastage_other"], 'indent': 2, 'bold': True},
        {'label': ''},
        {'label': 'TOTAL COGS', 'value': metrics["cogs"]["total_cogs"], 'bold': True, 'fill': total_fill}
    ])
    row = add_section("C. COST OF GOODS SOLD (COGS)", cogs_items, row)

    # D. Gross Profit
    gross_profit = metrics["revenue"]["net_revenue"] - metrics["cogs"]["total_cogs"]
    gross_margin = (gross_profit / metrics["revenue"]["net_revenue"] * 100) if metrics["revenue"]["net_revenue"] > 0 else 0
    gp_items = [
        {'label': 'Gross Profit', 'value': gross_profit, 'bold': True},
        {'label': f'Gross Margin %', 'value': f'{gross_margin:.1f}%', 'bold': True}
    ]
    row = add_section("D. GROSS PROFIT", gp_items, row)

    # E. OPEX
    opex_items = [
        {'label': 'E1. Labour & HR Costs', 'bold': True},
        {'label': f'Staff Salaries ({metrics["labour"]["employee_count"]} employees)', 'value': metrics["labour"]["salaries"], 'indent': 2},
        {'label': 'PF/ESIC (13.75%)', 'value': metrics["labour"]["pf_esic"], 'indent': 2},
        {'label': 'Overtime Allowance', 'value': metrics["labour"]["overtime"], 'indent': 2},
        {'label': 'Sub-Total: Labour', 'value': metrics["labour"]["total_labour"], 'indent': 2, 'bold': True},
        {'label': ''}
    ]

    # Add remaining OPEX sections similarly...
    opex_items.append({'label': 'E2. Occupancy & Utilities', 'bold': True})
    for key, value in metrics["opex"]["occupancy_details"].items():
        opex_items.append({'label': key.replace('_', ' ').title(), 'value': value, 'indent': 2})
    opex_items.append({'label': 'Sub-Total: Occupancy', 'value': metrics["opex"]["occupancy"], 'indent': 2, 'bold': True})
    opex_items.append({'label': ''})

    opex_items.append({'label': 'E3. Technology & Software', 'bold': True})
    for key, value in metrics["opex"]["technology_details"].items():
        opex_items.append({'label': key.replace('_', ' ').title(), 'value': value, 'indent': 2})
    opex_items.append({'label': 'Sub-Total: Technology', 'value': metrics["opex"]["technology"], 'indent': 2, 'bold': True})
    opex_items.append({'label': ''})

    opex_items.append({'label': 'E4. Sales & Marketing', 'bold': True})
    for key, value in metrics["opex"]["marketing_details"].items():
        opex_items.append({'label': key.replace('_', ' ').title(), 'value': value, 'indent': 2})
    opex_items.append({'label': 'Sub-Total: Marketing', 'value': metrics["opex"]["marketing"], 'indent': 2, 'bold': True})
    opex_items.append({'label': ''})

    opex_items.append({'label': 'E5. General & Administrative', 'bold': True})
    for key, value in metrics["opex"]["admin_details"].items():
        opex_items.append({'label': key.replace('_', ' ').title(), 'value': value, 'indent': 2})
    opex_items.append({'label': 'Sub-Total: General & Admin', 'value': metrics["opex"]["general_admin"], 'indent': 2, 'bold': True})
    opex_items.append({'label': ''})

    total_opex = (metrics['labour']['total_labour'] + metrics['opex']['occupancy'] +
                  metrics['opex']['technology'] + metrics['opex']['marketing'] +
                  metrics['opex']['general_admin'])
    opex_items.append({'label': 'TOTAL OPEX', 'value': total_opex, 'bold': True, 'fill': total_fill})
    row = add_section("E. OPERATING EXPENSES (OPEX)", opex_items, row)

    # F-M sections...
    ebitda = gross_profit - total_opex
    ebitda_margin = (ebitda / metrics["revenue"]["net_revenue"] * 100) if metrics["revenue"]["net_revenue"] > 0 else 0
    ebitda_items = [
        {'label': 'EBITDA', 'value': ebitda, 'bold': True},
        {'label': 'EBITDA Margin %', 'value': f'{ebitda_margin:.1f}%', 'bold': True}
    ]
    row = add_section("F. EBITDA", ebitda_items, row)

    da_items = [
        {'label': 'Equipment Depreciation', 'value': metrics["depreciation"]["equipment"], 'indent': 1},
        {'label': 'Brand Amortization', 'value': metrics["depreciation"]["brand"], 'indent': 1},
        {'label': 'Total D&A', 'value': metrics["depreciation"]["total"], 'bold': True}
    ]
    row = add_section("G. DEPRECIATION & AMORTIZATION", da_items, row)

    ebit = ebitda - metrics['depreciation']['total']
    ebit_items = [{'label': 'EBIT', 'value': ebit, 'bold': True}]
    row = add_section("H. EBIT", ebit_items, row)

    finance_items = [
        {'label': 'Interest on Loans', 'value': metrics["finance"]["interest"], 'indent': 1},
        {'label': 'Bank Charges', 'value': metrics["finance"]["bank_charges"], 'indent': 1},
        {'label': 'Total Finance Costs', 'value': metrics["finance"]["total"], 'bold': True}
    ]
    row = add_section("I. FINANCE COSTS", finance_items, row)

    pbt = ebit - metrics['finance']['total']
    pbt_items = [{'label': 'PBT', 'value': pbt, 'bold': True}]
    row = add_section("J. PROFIT BEFORE TAX (PBT)", pbt_items, row)

    tax_items = [{'label': 'Income Tax (26%)', 'value': metrics['tax'], 'bold': True}]
    row = add_section("K. INCOME TAX", tax_items, row)

    pat = pbt - metrics['tax']
    pat_margin = (pat / metrics["revenue"]["net_revenue"] * 100) if metrics["revenue"]["net_revenue"] > 0 else 0
    pat_items = [
        {'label': 'PAT', 'value': pat, 'bold': True, 'fill': total_fill},
        {'label': 'PAT Margin %', 'value': f'{pat_margin:.1f}%', 'bold': True, 'fill': total_fill}
    ]
    row = add_section("L. PROFIT AFTER TAX (PAT) ★", pat_items, row)

    aov = metrics["revenue"]["net_gmv"] / metrics["revenue"]["total_orders"] if metrics["revenue"]["total_orders"] > 0 else 0
    total_commission = metrics['revenue']['zomato_commission'] + metrics['revenue']['swiggy_commission']

    kpi_items = [
        {'label': 'Total Orders', 'value': metrics['revenue']['total_orders'], 'indent': 1},
        {'label': 'Average Order Value', 'value': aov, 'indent': 1},
        {'label': 'Food Cost % (of Net Revenue)', 'value': f"{(metrics['cogs']['total_raw_material']/metrics['revenue']['net_revenue']*100) if metrics['revenue']['net_revenue'] > 0 else 0:.1f}%", 'indent': 1},
        {'label': 'Labour Cost % (of Net Revenue)', 'value': f"{(metrics['labour']['total_labour']/metrics['revenue']['net_revenue']*100) if metrics['revenue']['net_revenue'] > 0 else 0:.1f}%", 'indent': 1},
        {'label': 'Platform Commission % (of GMV)', 'value': f"{(total_commission/metrics['revenue']['gross_gmv']*100) if metrics['revenue']['gross_gmv'] > 0 else 0:.1f}%", 'indent': 1}
    ]
    row = add_section("M. KEY PERFORMANCE INDICATORS", kpi_items, row)

    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20

    # Save
    wb.save(output_path)
    return output_path


def generate_pnl(start_date: str, end_date: str, restaurant_id: str, output_format: str = 'text'):
    """Main P&L generation function"""

    # Parse dates
    start_dt, end_dt, num_days = parse_date_range(start_date, end_date)

    # Connect to MongoDB
    db = get_mongodb_connection()

    # If the requested range extends beyond the dataset, clamp to the latest order date.
    start_dt, end_dt, adjusted = clamp_date_range_to_available_orders(db, start_dt, end_dt)
    num_days = (end_dt.date() - start_dt.date()).days + 1
    if adjusted:
        # Keep the report period consistent with the data actually used
        start_date = start_dt.date().isoformat()
        end_date = end_dt.date().isoformat()

    # Get settings
    settings = get_restaurant_settings(db, restaurant_id)

    # Calculate all metrics
    print("Calculating revenue breakdown...")
    revenue = calculate_revenue_breakdown(db, restaurant_id, start_dt, end_dt, settings)

    print("Calculating COGS...")
    cogs = calculate_cogs(db, restaurant_id, start_dt, end_dt)

    print("Calculating labour costs...")
    labour = calculate_labour_costs(db, settings, num_days)

    print("Calculating OPEX...")
    opex = calculate_opex(settings, num_days)

    print("Calculating depreciation...")
    depreciation = calculate_depreciation(settings, num_days)

    print("Calculating finance costs...")
    finance = calculate_finance_costs(settings, num_days)

    # Calculate PBT
    gross_profit = revenue["net_revenue"] - cogs["total_cogs"]
    total_opex = (labour["total_labour"] + opex["occupancy"] + opex["technology"] +
                  opex["marketing"] + opex["general_admin"])
    ebitda = gross_profit - total_opex
    ebit = ebitda - depreciation["total"]
    pbt = ebit - finance["total"]

    print("Calculating tax...")
    tax = calculate_tax(pbt, settings)

    # Compile metrics
    metrics = {
        "revenue": revenue,
        "cogs": cogs,
        "labour": labour,
        "opex": opex,
        "depreciation": depreciation,
        "finance": finance,
        "tax": tax
    }

    # Generate report
    if output_format == 'text':
        period_note = ""
        if adjusted:
            period_note = (
                "NOTE: Requested date range was adjusted to match the latest available "
                "order data in MongoDB."
            )
        if revenue.get("total_orders", 0) == 0 and revenue.get("gross_gmv", 0) == 0:
            try:
                n = db.orders.estimated_document_count()
                if n > 0:
                    period_note = (
                        "NOTE: No orders were found in this date range, so revenue is ₹0. "
                        "Try selecting a different date range."
                    )
            except Exception:
                pass
        text_report = generate_text_report(metrics, start_date, end_date, period_note)
        print("\n" + text_report)
        print(f"\nSUCCESS:Detailed P&L generated for {revenue['total_orders']} orders")
        return 0
    elif output_format == 'excel':
        # Create output directory if it doesn't exist (use REPORTS_DIR from env for local/Docker)
        default_reports = "/app/backend/static/reports"
        output_dir = Path(os.getenv("REPORTS_DIR", default_reports))
        output_dir.mkdir(parents=True, exist_ok=True)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"PnL_{start_date}_{end_date}_{timestamp}.xlsx"
        output_path = output_dir / filename

        print("Generating Excel report...")
        generate_excel_report(metrics, start_date, end_date, str(output_path))

        print(f"\nSUCCESS:Excel report saved to {output_path}")
        print(f"FILE_PATH:{output_path}")  # For chatbot to parse
        return 0
    else:
        print(f"ERROR:Unsupported format '{output_format}'. Use 'text' or 'excel'")
        sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) < 3 or len(sys.argv) > 5:
        print("ERROR:Usage: python generate_pnl.py <start_date> <end_date> [format] [restaurant_id]")
        sys.exit(1)

    start_date = sys.argv[1]
    end_date = sys.argv[2]
    output_format = sys.argv[3] if len(sys.argv) >= 4 else 'text'
    restaurant_id = sys.argv[4] if len(sys.argv) >= 5 else os.getenv('RESTAURANT_ID', 'default')

    sys.exit(generate_pnl(start_date, end_date, restaurant_id, output_format))
